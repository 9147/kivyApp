import json
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import requests
import os
import math

url = "https://0000manoj0000.pythonanywhere.com/"
# url = 'http://127.0.0.1:8000/'

def store_token(token):
    with open('token.json', 'w') as f:
        json.dump({'token': token}, f)


def get_stored_token():
    try:
        with open('token.json', 'r') as f:
            data = json.load(f)
            return data.get('token')
    except FileNotFoundError:
        return ""

def get_common_text(text1, text2):
    text1 = text1.split(' ')
    text2 = text2.split(' ')
    common_text = []
    for t1, t2 in zip(text1, text2):
        if t1 == t2:
            common_text.append(t1)
        else:
            break
    return ' '.join(common_text).strip('term').strip()

# get the scheme from the server and store it if its updated
def get_data_scheme():
    global url
    data = {"request": "scheme"}
    headers = {'Authorization': f'Token {get_stored_token()}'}
    response = requests.post(url + "data/", headers=headers, data=data)
    if response.status_code == 200:
        response_json = response.json()
        new_scheme = {"classes": response_json.get("classes")}
        if os.path.exists('scheme.json'):
            with open('scheme.json', 'r') as f:
                old_scheme = json.load(f)
            if old_scheme != new_scheme:
                print("Database scheme has changed.")
                update_database(new_scheme)
            else:
                print("Database scheme is up-to-date.")
        else:
            with open('scheme.json', 'w') as f:
                json.dump(new_scheme, f)
            create_database()
    else:
        print("Failed to get the scheme")

    



def update_database(new_scheme):
    print("Updating the database")
    with open('scheme.json', 'w') as f:
        json.dump(new_scheme, f)

def change_text_field(instance):
    text_values = ['Beginner', 'Progressing', 'Proficient']
    if instance.text in text_values:
        index = text_values.index(instance.text)
        if index == len(text_values) - 1:
            instance.text = text_values[0]
        else:
            instance.text = text_values[index + 1]
    else:
        instance.text = text_values[0]

def create_database():
    print("Creating the database")
    with open('scheme.json', 'r') as f:
        scheme = json.load(f)
        classes = scheme['classes']
        for cls in classes:
            key = cls['name']
            # check if file exists
            file_path = f'resources' + '/' + key + '.xlsx'
            if os.path.exists(file_path):
                workbook = load_workbook(filename=file_path)
            else:
                workbook = Workbook()
                workbook.save(file_path)
            values = ['cover_page', 'first_page','Image_page','development_page','feedback_page']
            for value in values:
                if value in ['cover_page','first_page']:
                    worksheet = workbook[value] if value in workbook else None
                    # If the worksheet does not exist, create a new one
                    if worksheet is None:
                        worksheet = workbook.create_sheet(value)
                    sub = cls[value]['report_fields']
                    i = 0
                    for s in sub:
                        worksheet.cell(row=1, column=i + 1, value=s['name'])
                        i += 1
                elif value=='development_page':
                    sub = cls[value]
                    for s in sub:
                        worksheet = workbook[s['development_goal']] if s['development_goal'] in workbook else None
                        # If the worksheet does not exist, create a new one
                        if worksheet is None:
                            worksheet = workbook.create_sheet(s['development_goal'])
                        sub_li=s['sections']
                        i = 0
                        for sub in sub_li:
                            li=sub['learning_outcome']
                            for l in li:
                                worksheet.cell(row=1, column=i + 1, value=l['code']+" term 1")
                                worksheet.cell(row=1, column=i+2, value=l['code']+" term 2")
                                i += 2
                elif value=='feedback_page':
                    sub=cls[value]["sections"]
                    print(sub)
                    worksheet = workbook[value] if value in workbook else None
                    if worksheet is None:
                        worksheet = workbook.create_sheet(value)
                    i=0
                    for s in sub:
                        sub_li=s["Fields"]
                        print(sub_li)
                        for l in sub_li:
                            print(l['name'])
                            worksheet.cell(row=1, column=i + 1, value=l['name'])
                            i += 1
                elif value=='Image_page':
                    sub=cls[value]["images"]
                    worksheet = workbook[value] if value in workbook else None
                    if worksheet is None:
                        worksheet = workbook.create_sheet(value)
                    # worksheet.cell(row=1, column=1, value='Image')
                    for i in range(len(sub)):
                        worksheet.cell(row=1, column=i + 1, value=sub[i]['title'])
            workbook.save(file_path)

# gets the height of the string in pdf
def get_height(pdf, text, width):
    word_length = pdf.get_string_width(text+"m")
    lines = math.ceil(word_length / width)
    # print(word_length,text, lines)
    return lines

# check if the value is a path
def check_if_path(path):
    return os.path.exists(path) and os.path.isfile(path)


def get_username():
    with open('user.json', 'r') as f:
        data = json.load(f)
        return data.get('username')