import math
import os
import shutil
import threading
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.widget import Widget
from kivymd.app import MDApp
from kivy.uix.screenmanager import ScreenManager, Screen
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.label import MDLabel
from kivymd.toast import toast
from kivymd.uix.snackbar import MDSnackbar, MDSnackbarActionButton
from kivymd.uix.button import MDFlatButton
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton
from kivymd.uix.button import MDRectangleFlatIconButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.list import OneLineAvatarIconListItem
import glob
import requests
from kivymd.uix.filemanager import MDFileManager
import json
from openpyxl import load_workbook
from kivy.uix.scrollview import ScrollView
from kivymd.uix.textfield import MDTextField
from fpdf import FPDF
from networking import get_global_ipv6_address, start_server, connect_to_server_thread
from kivy.clock import Clock
from imageConversion import encode_image_to_base64
import os
from dependant import *


Builder.load_string('''
#:import toast kivymd.toast.toast
                    
<ItemConfirm>
    on_release: root.set_icon(check)

    CheckboxLeftWidget:
        id: check
        group: "check"

<LoginScreen>:
    orientation: "vertical"
    padding: "10dp"
    spacing: "10dp"

    MDTopAppBar:
        title: "ReportGen"
        pos_hint: {"top": 1}

    Widget:

    BoxLayout:
        size_hint: (None, None)
        size: ("300dp", "200dp")
        pos_hint: {"center_x": 0.5, "center_y": 0.5}
        orientation: "vertical"
        spacing: "10dp"

        MDTextField:
            id: username
            hint_text: "username"
            helper_text: "Enter your esername"
            helper_text_mode: "on_focus"
            icon_right: "account"
            size_hint: (.8,None)
            icon_right_color: app.theme_cls.primary_color
            pos_hint: {"center_x": 0.5, "center_y": 1}

        MDTextField:
            id: password
            hint_text: "Password"
            helper_text: "Enter your password"
            helper_text_mode: "on_focus"
            password: True
            size_hint: (.8,None)
            icon_right: "eye-off"
            icon_right_color: app.theme_cls.primary_color
            pos_hint: {"center_x": 0.5, "center_y": 0.5}

        MDRaisedButton:
            text: "Login"
            pos_hint: {"center_x": 0.5, "center_y": 0.5}
            on_release: app.on_login(username.text, password.text)

    Widget:

<HomeScreen>:
    name: 'home'
    MDTopAppBar:
        title: "ReportGen"
        pos_hint: {"top": 1}
        right_action_items: [["logout", lambda x: app.logout()]]


    BoxLayout:
        size_hint: (None, None)
        pos_hint: {"center_x": 0.5, "center_y": 0.5}
        orientation: "vertical"
        spacing: "30dp"

        Widget:

        MDRaisedButton:
            text: "Edit Report"
            size_hint: (None, None)
            size: ("400dp", "150dp")
            pos_hint: {"center_x": 0.5}
            on_release: app.root.current = 'edit'

        MDRaisedButton:
            text: "Print Report"
            size_hint: (None, None)
            size: ("400dp", "150dp")
            pos_hint: {"center_x": 0.5}
            on_release: app.root.current = 'print'

        MDRaisedButton:
            text: "Add Report"
            size_hint: (None, None)
            size: ("400dp", "150dp")
            pos_hint: {"center_x": 0.5}
            on_release: app.root.current = 'add'

        Widget:

<EditScreen>:
    name: 'edit'
    MDTopAppBar:
        title: "ReportGen"
        pos_hint: {"top": 1}
        right_action_items: [["home", lambda x: app.go_home()], ["logout", lambda x: app.logout()]]

    BoxLayout:
        id: container
        orientation: 'vertical'
        size_hint: (1, 1)
        pos_hint: {"center_x": 0.5, "center_y": 0.5}


<PrintScreen>:
    name: 'print'
    MDTopAppBar:
        title: "ReportGen"
        pos_hint: {"top": 1}
        right_action_items: [["home", lambda x: app.go_home()], ["logout", lambda x: app.logout()]]

    BoxLayout:
        id: container
        orientation: 'vertical'
        size_hint: (1, 1)
        pos_hint: {"center_x": 0.5, "center_y": 0.5}

<AddScreen>:
    name: 'add'
    MDTopAppBar:
        title: "ReportGen"
        pos_hint: {"top": 1}
        right_action_items: [["home", lambda x: app.go_home()], ["logout", lambda x: app.logout()]]

    BoxLayout:
        id: container
        orientation: 'vertical'
        size_hint: (1, 1)
        pos_hint: {"center_x": 0.5, "center_y": 0.5}

''')




def ImLive(dt):
    headers = {'Authorization': f'Token {get_stored_token()}'}
    if headers['Authorization']=="":
        toast("Your session has expired")
        MainApp().logout()
    new_url = url + 'live/'
    data={'request':'live','ipv6':get_global_ipv6_address()}
    # print(data)
    try:
        response = requests.post(new_url, headers=headers,data=data)
        if not response.status_code == 200:
            toast("Connection lost!!")
    except requests.exceptions.RequestException as e:
        toast("Server connection error")
    print("I'm live!")

class CustomFileManager(MDFileManager):
    def show(self, path):
        '''Forms the body of a directory. Called when opening a directory.'''

        self.current_path = path
        super().show(path)
        self.selection_button.opacity = 0  # hide the selection_button
        self.selection_button.disabled = True 


class ImageManager():
    def __init__(self):
        Window.bind(on_keyboard=self.events)
        self.manager_open = False
        self.instance = None
        self.file_manager = CustomFileManager(
            exit_manager=self.exit_manager, 
            select_path=self.select_path,
            ext=['.jpg', '.png', '.jpeg'],# only show these types of files
            preview=True, # allow preview of images
            icon_selection_button="none",
        )

    def file_manager_open(self,instance):
        print(instance)
        self.instance=instance
        self.file_manager.show(os.path.expanduser("~"))  # output manager to the screen
        self.manager_open = True

    def select_path(self, path: str):
        '''
        It will be called when you click on the file name
        or the catalog selection button.

        :param path: path to the selected directory or file;
        '''
        # get file name
        self.instance.text=path
        self.exit_manager()
        self.instance= None
        toast(path)

    def exit_manager(self, *args):
        '''Called when the user reaches the root of the directory tree.'''
        self.instance=None
        self.manager_open = False
        self.file_manager.close()

    def events(self, instance, keyboard, keycode, text, modifiers):
        '''Called when buttons are pressed on the mobile device.'''

        if keyboard in (1001, 27):
            if self.manager_open:
                self.file_manager.back()
        return True
    

class SaveImage:
    def __init__(self):
        pass

    def addImage(self, index, from_path, to_path):
        if index == 0:
            self.MeImage = [from_path, to_path]
        if index == 1:
            self.FamilyImage = [from_path, to_path]

    def save_image(self):
        if self.MeImage[0]!=self.MeImage[1]:
            shutil.copy(self.MeImage[0], self.MeImage[1])
        if self.FamilyImage[0]!=self.FamilyImage[1]:
            shutil.copy(self.FamilyImage[0], self.FamilyImage[1])


class ItemConfirm(OneLineAvatarIconListItem):
    divider = None

    def on_release(self):
        if self.text.startswith("Other"):
            self.dialog = MDDialog(
                title="Enter your text",
                type="custom",
                content_cls=MDTextField(),
                buttons=[
                    MDFlatButton(
                        text="CANCEL",
                        theme_text_color="Custom",
                        text_color=self.theme_cls.primary_color,
                        on_release=self.close_dialog,
                    ),
                    MDFlatButton(
                        text="OK",
                        theme_text_color="Custom",
                        text_color=self.theme_cls.primary_color,
                        on_release=self.close_dialog_with_input,
                    ),
                ],
            )
            self.dialog.open()
        else:
            super().on_release()

    def close_dialog(self, *args):
        self.dialog.dismiss()
    
    def set_icon(self, icon_name):
        # change icon of the checkbox
        self.ids.check.active = True

    def close_dialog_with_input(self, *args):
        # change this text of other to input text
        self.text = "Other ("+self.dialog.content_cls.text+")"
        self.dialog.dismiss()


class LoginScreen(Screen, MDBoxLayout):
    pass


class HomeScreen(Screen):

    def on_enter(self, *args):
        # check if user.json exists
        if os.path.exists('user.json'):
            with open('user.json') as f:
                user = json.load(f)
                if user.get('username', None) is None:
                    self.manager.current = 'login'
        else:
            self.manager.current = 'login'


        # open user.json file
        access_files = {}
        with open('scheme.json') as f:
            data = json.load(f)
            for d in data['classes']:
                access_files[d['name']]=d['commit_number']
        with open('user.json') as f:
            user = json.load(f)
            if user.get('commit_no', None) is None:
                user['commit_no'] = {}
            for file in access_files:
                if user['commit_no'].get(file, None) is None:
                    user['commit_no'][file] = 0
                if user['commit_no'][file] < access_files[file]:
                    toast(f"You are lacking by {access_files[file]-user['commit_no'][file]} commits for class {file}")
                    

            with open('user.json', 'w') as f:
                json.dump(user, f)
        get_data_scheme()


class EditScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.file_manager = ImageManager()
        self.selected_row = None
        self.workbook_active = None
        self.workbook = None
        self.newImage=None
        self.sheets = None
        self.dialog = None
        self.current_sheet_index = 0
        self.values = []
        self.development_page_count = 0
        global url
        self.url = url
        self.section_no=""

    def on_enter(self, *args):
        self.values = []
        self.ids.container.clear_widgets()
        self.get_xlsx_files()

    def get_xlsx_files(self):
        files = glob.glob('resources/*.xlsx')
        dropdown_items = [os.path.basename(f) for f in files]
        self.add_buttons(dropdown_items)

    def add_buttons(self, files):
        widge = Widget()
        self.ids.container.add_widget(widge)
        for file in files:
            button = MDRaisedButton(text=file, on_release=self.on_button_click, pos_hint={"center_x": 0.5})
            self.ids.container.add_widget(button)
        widge = Widget()
        self.ids.container.add_widget(widge)

    def on_button_click(self, instance):
        self.workbook_active = instance.text
        workbook = load_workbook(filename=f'resources/{instance.text}')
        # get the first sheet
        sheets = [sheet.title for sheet in workbook.worksheets]

        # Create a list to hold the sheets
        self.sheets = sheets
        self.workbook = workbook
        self.current_sheet_index = 0
        self.ids.container.clear_widgets()
        self.ids.container.add_widget(Widget())
        # add a text field for the admission number
        text_field = MDTextField(hint_text='Admission Number', pos_hint={"center_x": 0.5}, id='admissionNo')
        self.ids.container.add_widget(text_field)
        self.ids.container.add_widget(
            MDRaisedButton(text='Match', on_release=self.match_admissionNo, pos_hint={"center_x": 0.5}))
        self.ids.container.add_widget(Widget())

    def match_admissionNo(self, instance):
        match=False
        for widget in self.ids.container.children:
            if isinstance(widget, MDTextField) and widget.id == 'admissionNo':
                text_field = widget
                break
        admissionNo = text_field.text
        self.admission_number=str(admissionNo).strip()
        if admissionNo:
            sheet = self.workbook['cover_page']
            
    #         in first row find the cell with value admission number
            for cell in sheet[1]:
                if cell.value == 'Admission Number':
                    row:int=2
                    while row <= sheet.max_row:
                        if str(sheet.cell(row=row, column=cell.column).value).strip() == str(admissionNo).strip():
                            self.on_admissionNo_match()
                            match=True
                            self.selected_row = row
                        row += 1
        if not match:
            snackbar = MDSnackbar(
                MDLabel(
                    text="Admission Number does not match",
                ),
                MDSnackbarActionButton(
                    text="DISMISS",
                    on_release=lambda *args: snackbar.dismiss(),
                    theme_text_color="Custom",
                    text_color="#8E353C",
                ),
                y=dp(24),
                pos_hint={"center_x": 0.5},
                size_hint_x=0.5,
                md_bg_color="#E8D8D7",
            )
            snackbar.open()


    def on_admissionNo_match(self):
        # remover everything inside the container
        self.ids.container.clear_widgets()
        # Create a ScrollView
        scroll_view = ScrollView(do_scroll_x=False, size_hint=(None, None), size=(400, 450), pos_hint={"center_x": 0.5})
        # Create a BoxLayout inside the ScrollView
        box_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        box_layout.bind(minimum_height=box_layout.setter('height'))
        scroll_view.add_widget(box_layout)

        # Create MDTextField for each cell in the first row of the first sheet
        self.create_text_fields(box_layout, self.current_sheet_index)
        self.development_page_count = 0

        # Add the ScrollView to the container
        self.ids.container.add_widget(Widget())
        self.ids.container.add_widget(scroll_view)
        self.ids.container.add_widget(Widget())


    def create_button_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

                # Get the first row
        row = worksheet[1]
        val = []

        # Create a button for each cell in the row
        for i in range(0, len(row), 2):
            cell1 = row[i]
            cell2 = row[i+1]
            if cell1.value and cell2.value:
                common_text = get_common_text(cell1.value, cell2.value)
                component = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 60))
                label = MDLabel(text=common_text)
                val1=str(worksheet.cell(row=self.selected_row, column=cell1.column).value)
                val2=str(worksheet.cell(row=self.selected_row, column=cell2.column).value)
                button1 = MDRaisedButton(text="Term 1" if val1==str(None) else val1, on_press=lambda x: change_text_field(x),id=common_text+'1')
                button2 = MDRaisedButton(text="Term 2" if val2==str(None) else val2,on_press=lambda x: change_text_field(x),id=common_text+'2')
                component.add_widget(label)
                component.add_widget(button1)
                component.add_widget(button2)
                box_layout.add_widget(component)


        # # Create a MDTextField for each cell in the row
        # for cell in row:
        #     if cell.value:
        #         text_field = MDTextField(hint_text=str(cell.value), text=str(worksheet.cell(row=self.selected_row, column=cell.column).value) if worksheet.cell(row=self.selected_row, column=cell.column).value else "")
        #         box_layout.add_widget(text_field)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def create_feedback_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]

        # Create a button for each cell in the row
        for i in range(0, len(row)):
                component = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 60))
                label = MDLabel(text=row[i].value)
                val=str(worksheet.cell(row=self.selected_row, column=row[i].column).value)
                # a button with dropdown arrow on the right

                button1 = MDRectangleFlatIconButton(text="None" if val==str(None) else val,icon="menu-down", on_press=lambda x:self.show_confirmation_dialog(x),id=row[i].value)
                component.add_widget(label)
                component.add_widget(button1)
                box_layout.add_widget(component)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def create_text_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]
        val = []

        # Create a MDTextField for each cell in the row
        for cell in row:
            if cell.value:
                if cell.value.strip()=='Admission Number':
                    text_field = MDTextField(hint_text=str(cell.value), text=str(worksheet.cell(row=self.selected_row, column=cell.column).value) if worksheet.cell(row=self.selected_row, column=cell.column).value else "",readonly=True )
                else:
                    text_field = MDTextField(hint_text=str(cell.value), text=str(worksheet.cell(row=self.selected_row, column=cell.column).value) if worksheet.cell(row=self.selected_row, column=cell.column).value else "")
                box_layout.add_widget(text_field)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def show_confirmation_dialog(self,instance):
        feedbacks_options=[]
        with open('scheme.json') as f:
            data = json.load(f)
            for d in data['classes']:
                if d['name'] == self.workbook_active.split('.')[0]:
                    data = d
                    break
            data=data['feedback_page']['sections']
            for d in data:
                for d1 in d['Fields']:
                    if d1['name'] == instance.id:
                        feedbacks_options = d1["options"]
                        break

        item=[ItemConfirm(text=i['choice']) for i in feedbacks_options]
        item.append(ItemConfirm(text='Other'))
        # print(feedbacks_options)
        self.dialog = MDDialog(
            title=instance.id,
            # feedbacks=feedbacks,
            type="confirmation",
            items=item,
            buttons=[
                MDFlatButton(
                    text="CANCEL",
                    theme_text_color="Custom",
                    on_release=self.close_dialog,
                ),
                MDFlatButton(
                    text="OK",
                    theme_text_color="Custom",
                    on_release=lambda x:self.close_dialog(x, True,instance),
                ),
            ],
        )
        self.dialog.open()

    def close_dialog(self, inst, update=False,instance=None):
        if update:
            val=[i.text for i in self.dialog.items if i.ids.check.active]
            # updated the text of button that has colled the function
            if len(val)>0:
                instance.text=val[0]
        self.dialog.dismiss()

    def create_image_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]

        # Create a button for each cell in the row
        for cell in row:
            if cell.value:
                component = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 60))
                label = MDLabel(text=cell.value)
                val=str(worksheet.cell(row=self.selected_row, column=cell.column).value)
                # a button with dropdown arrow on the right
                button1 = MDRectangleFlatIconButton(text="None" if val==str(None) else val,id=cell.value,icon="folder")
                button1.bind( on_press=lambda x,button1=button1:self.file_manager.file_manager_open(button1))
                component.add_widget(label)
                component.add_widget(button1)
                box_layout.add_widget(component)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def on_next_button_click(self, instance):
        # iterate throught the text fields and get the values
        value_li = []
        # get the current sheet name
        sheet = self.sheets[self.current_sheet_index]
        if sheet in ['cover_page','first_page']:
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for text_field in widget.children:
                                if isinstance(text_field, MDTextField):
                                    print(text_field.text)
                                    value_li.append(text_field.text)
        elif self.sheets[self.current_sheet_index] in ['feedback_page']:
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for boxs in widget.children:
                                for button in boxs.children:
                                    if isinstance(button, MDRectangleFlatIconButton):
                                        value_li.append(button.text)
        elif self.sheets[self.current_sheet_index] in ['Image_page']:
            # Check if the images folder exists
            if not os.path.exists('resources/images'):
                os.makedirs('resources/images')
            print("values: ",self.values[1][1])
            if not os.path.exists(f'resources/images/{self.values[1][1]}'):
                os.makedirs(f'resources/images/{self.values[1][1]}')
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    i=0
                    self.newImage = SaveImage()
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for boxs in  widget.children:
                                files=['Me','Family']
                                for button in boxs.children:
                                    if isinstance(button, MDRectangleFlatIconButton):
                                            # print(i,button.text)
                                            ext=button.text.split('.')[-1]
                                            # print("ext: ",ext)
                                            value_li.append(f'resources/images/{self.values[1][1]}/{files[i]}.{ext}')
                                            self.newImage.addImage(i,button.text,f'resources/images/{self.values[1][1]}/{files[i]}.{ext}')
                                            # print(files[i])
                                            i+=1
        else:
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for boxs in widget.children:
                                for button in boxs.children:
                                    if isinstance(button, MDRaisedButton):
                                        if button.text == 'Term 1' or button.text == 'Term 2':
                                            value_li.append("")
                                        else:
                                            value_li.append(button.text)
        # print(value_li[::-1])

        # Increment the current sheet index
        self.current_sheet_index += 1
        self.values.append(value_li[::-1])

        # Clear the container
        self.ids.container.clear_widgets()
        # Create a ScrollView
        scroll_view = ScrollView(do_scroll_x=False, size_hint=(None, None), size=(400, 450), pos_hint={"center_x": 0.5})
        # Create a BoxLayout inside the ScrollView
        box_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        box_layout.bind(minimum_height=box_layout.setter('height'))
        scroll_view.add_widget(box_layout)

        click_next_button = False

        # If there are more sheets, create text fields for the next sheet
        if self.current_sheet_index < len(self.sheets):
            with open('scheme.json') as f:
                data = json.load(f)
                for d in data['classes']:
                    if d['name'] == self.workbook_active.split('.')[0]:
                        data = d
                        break
                else:
                    # notify that there was an error and returm to the main screen
                    toast('An error occured please contact the developer! This is related to your access to this section')
                    self.manager.current = 'home'
            # access_accounts=data[self.sheets[self.current_sheet_index]+"_access"]
            # print("access accounts",access_accounts)
            if self.sheets[self.current_sheet_index] in ['cover_page','first_page']:
                access_accounts=data[self.sheets[self.current_sheet_index]+"_access"]
                self.create_text_fields(box_layout, self.current_sheet_index)
                # get dict from scheme.json
                # if access_accounts
                if not (access_accounts['username']==get_username() or get_username() == 'admin'):
                    click_next_button = True
                else:
                    self.section_no+=','+str(self.current_sheet_index)
            elif self.sheets[self.current_sheet_index] in ['Image_page']:
                access_accounts=data[self.sheets[self.current_sheet_index]+"_access"]
                self.create_image_fields(box_layout, self.current_sheet_index)
                if not (access_accounts['username']==get_username() or get_username() == 'admin'):
                    click_next_button=True
                else:
                    self.section_no+=','+str(self.current_sheet_index)
            elif self.sheets[self.current_sheet_index] in ['feedback_page']:
                self.section_no+=','+str(self.current_sheet_index)
                self.create_feedback_fields(box_layout, self.current_sheet_index)
            else:
                self.create_button_fields(box_layout, self.current_sheet_index)
                access_accounts=[dic['username'] for dic in data['development_page_access'][self.development_page_count]['Auth_teachers_access']]
                # print("access accounts",access_accounts)
                self.development_page_count += 1
                if not (get_username() in access_accounts or get_username() == 'admin'):
                    click_next_button=True
                else:
                    self.section_no+=','+str(self.current_sheet_index)
            self.ids.container.add_widget(Widget())
            self.ids.container.add_widget(scroll_view)
            self.ids.container.add_widget(Widget())
            if click_next_button:
                self.on_next_button_click(instance)
        else:
            #             display its done and add a home button
            self.ids.container.clear_widgets()
            self.ids.container.add_widget(Widget())
            self.ids.container.add_widget(MDLabel(text='Done', halign='center', theme_text_color='Primary'))
            self.ids.container.add_widget(Widget())
            for sheet, value in zip(self.sheets, self.values):
                worksheet = self.workbook[sheet]
            #      add values to the selected rows
                for val,a in zip(worksheet[self.selected_row],value):
                    val.value=a
            #     col=0
            #     for a in value:
            #         worksheet.cell(row=self.selected_row, column=col + 1, value=a)
            if self.newImage or self.newImage=='None':
                self.newImage.save_image()
            self.workbook.save(f'resources/{self.workbook_active}')
            url = self.url + "update/"
            print("value:",self.values[1])
            data = {"request": "update","section_no":self.section_no,'admission_no':self.values[1][1],'class_name':self.workbook_active.split('.')[0]}
            headers = {'Authorization': f'Token {get_stored_token()}'}
            response = requests.post(url, headers=headers, data=data)
            if response.status_code == 200:
                toast('Server notified successfully\n Your commit number is: '+str(response.json()['commit_no']))
                # open file user.json
                with open('user.json') as f:
                    user = json.load(f)
                    last_updated_commit_no = user.get("commit_no",0)
                    print(last_updated_commit_no)
                    if last_updated_commit_no==0:
                        user['commit_no']={self.workbook_active.split('.')[0]:0}
                        last_updated_commit_no = 0
                    else:
                        # print(self.workbook_active.split('.')[0])
                        last_updated_commit_no = last_updated_commit_no.get(self.workbook_active.split('.')[0],0)
                    # print("commit:",last_updated_commit_no)
                    if last_updated_commit_no + 1 == response.json()['commit_no']:
                        user['commit_no'][self.workbook_active.split('.')[0]]=response.json()['commit_no']
                        # update the user in the user.json file
                        with open('user.json', 'w') as f:
                            json.dump(user, f)
                print(response.json()['devices'])
                for device in response.json()['devices']:
                    # check that device ip is not loop back ip
                    ip_address=get_global_ipv6_address()
                    if not device['device_ip'].startswith('fe80') and not device['device_ip'].startswith('fd') and device['device_ip'] != '::1' and device['device_ip'] != ip_address:
                        sheets = [sheet.title for sheet in self.workbook.worksheets]
                        section_no=self.section_no.strip(',')
                        section_no=list(map(int,section_no.split(',')))
                        admission_no=self.values[1][1]
                        sheet=self.workbook['cover_page']
                        #in first row find the cell with value admission number
                        match=False
                        for cell in sheet[1]:
                            if cell.value == 'Admission Number':
                                row:int=2
                                while row <= sheet.max_row:
                                    if str(sheet.cell(row=row, column=cell.column).value).strip() == str(admission_no).strip():
                                        match=True
                                        selected_row = row
                                    row += 1
                        result= {}
                        files={}
                        if match:
                            for section in section_no:
                                for a in self.workbook[sheets[section]][selected_row]:
                                    if a.value and check_if_path(a.value):
                                        # result[section]='file_to_share'+str(file_count)
                                        files[a.value]=encode_image_to_base64(a.value)
                                result[section]=[a.value for a in self.workbook[sheets[section]][selected_row]]
                            print("results:",result)
                            # open file user.json
                            with open('user.json') as f:
                                user = json.load(f)
                                last_updated_commit_no = user.get("commit_no",0)
                                if last_updated_commit_no==0:
                                    user['commit_no']={self.workbook_active.split('.')[0]:0}
                                    last_updated_commit_no = 0
                                else:
                                    last_updated_commit_no = last_updated_commit_no.get(self.workbook_active.split('.')[0],0)
                                if last_updated_commit_no + 1 == response.json()['commit_no']:
                                    user['commit_no'][self.workbook_active.split('.'[0])]=response.json()['commit_no']
                                    # update the user in the user.json file
                                    with open('user.json', 'w') as f:
                                        json.dump(user, f)
                        json_data = {
                        "message": "Initiating commit push",
                        "commit_no": str(response.json()['commit_no']),
                        "admission_no": self.values[1][1],
                        "class_name": self.workbook_active.split('.')[0],
                        "section_no": self.section_no,
                        "results": result,
                        'files': files
                        }

                        # print(json_data)  # Debug print to inspect the JSON data
                        with open('data.json', 'w') as f:
                            json.dump(json_data, f)
                        connect_to_server_thread(device['device_ip'], 1680, json_data)
            else:
                # create a file named notification.txt
                with open('notification.txt', 'a') as f:
                    f.write("\n"+self.section_no+':'+self.values[1][1]+":"+self.workbook_active.split('.')[0])
                toast('An error occured while notifying the server')


class PrintScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.selected_row = None
        self.workbook_active = None
        self.workbook = None
        self.sheets = None
        self.current_sheet_index = 0
        self.values = []
        

    def on_enter(self, *args):
        self.values = []
        self.ids.container.clear_widgets()
        self.get_xlsx_files()

    def get_xlsx_files(self):
        files = glob.glob('resources/*.xlsx')
        dropdown_items = [os.path.basename(f) for f in files]
        self.add_buttons(dropdown_items)

    def add_buttons(self, files):
        widge = Widget()
        self.ids.container.add_widget(widge)
        for file in files:
            button = MDRaisedButton(text=file, on_release=self.on_button_click, pos_hint={"center_x": 0.5})
            self.ids.container.add_widget(button)
        widge = Widget()
        self.ids.container.add_widget(widge)

    def on_button_click(self, instance):
        self.workbook_active = instance.text
        workbook = load_workbook(filename=f'resources/{instance.text}')
        # get the first sheet
        sheets = [sheet.title for sheet in workbook.worksheets]

        # Create a list to hold the sheets
        self.sheets = sheets
        self.workbook = workbook
        self.current_sheet_index = 0
        self.ids.container.clear_widgets()
        self.ids.container.add_widget(Widget())
        # add a text field for the admission number
        text_field = MDTextField(hint_text='Admission Number', pos_hint={"center_x": 0.5}, id='admissionNo')
        self.ids.container.add_widget(text_field)
        self.ids.container.add_widget(
            MDRaisedButton(text='Match', on_release=self.match_admissionNo, pos_hint={"center_x": 0.5}))
        self.ids.container.add_widget(Widget())

    def match_admissionNo(self, instance):
        match=False
        for widget in self.ids.container.children:
            if isinstance(widget, MDTextField) and widget.id == 'admissionNo':
                text_field = widget
                break
        admissionNo = text_field.text
        if admissionNo:
            sheet = self.workbook['cover_page']
    #         in first row find the cell with value admission number
            for cell in sheet[1]:
                if cell.value == 'Admission Number':
                    row:int=2
                    while row <= sheet.max_row:
                        if str(sheet.cell(row=row, column=cell.column).value).strip() == str(admissionNo).strip():
                            match=True
                            self.selected_row = row
                        row += 1
        if not match:
            snackbar = MDSnackbar(
                MDLabel(
                    text="Admission Number does not match",
                ),
                MDSnackbarActionButton(
                    text="DISMISS",
                    on_release=lambda *args: snackbar.dismiss(),
                    theme_text_color="Custom",
                    text_color="#8E353C",
                ),
                y=dp(24),
                pos_hint={"center_x": 0.5},
                size_hint_x=0.5,
                md_bg_color="#E8D8D7",
            )
            snackbar.open()
        else:
            self.get_images()
            self.on_admissionNo_match()


    def on_admissionNo_match(self):
        # remover everything inside the container
        self.ids.container.clear_widgets()
        print('admission number matched')
        self.add_save_button()

    def add_save_button(self):
        button = MDRaisedButton(text='Save PDF', on_release=self.save_pdf, pos_hint={"center_x": 0.5})
        self.ids.container.add_widget(Widget())
        self.ids.container.add_widget(MDTextField(hint_text='Save name', id='save_location', pos_hint={"center_x": 0.5}))
        self.ids.container.add_widget(button)
        self.ids.container.add_widget(Widget())

    def save_pdf(self,instance):
    #    save the file to the location
        for widget in self.ids.container.children:
            if isinstance(widget, MDTextField) and widget.id == 'save_location':
                text_field = widget
                break
        try:
            save_location = "D:/Reports/"+self.workbook_active.split('.')[0]+"/"+text_field.text
            os.makedirs(os.path.dirname(save_location), exist_ok=True)  # Create directories if not present
            if not save_location.endswith('.pdf'):
                save_location+='.pdf'
            if save_location:
                shutil.move(f'resources/{self.workbook_active.split(".")[0]}.pdf', save_location)
                toast('PDF saved successfully')
                # display it is done
                self.ids.container.clear_widgets()
                self.ids.container.add_widget(Widget())
                self.ids.container.add_widget(MDLabel(text='PDF saved successfully', halign='center', theme_text_color='Primary'))
                self.ids.container.add_widget(Widget())
            else:
                toast('Please enter a save location')
        except FileNotFoundError:
            self.ids.container.clear_widgets()
            self.ids.container.add_widget(Widget())
            self.ids.container.add_widget(MDLabel(text='An unknown error occured please contact!', halign='center', theme_text_color='Primary'))
            self.ids.container.add_widget(Widget())
            toast('File not found')


    def get_images(self):
        # extract the file name from the workbook
        file_name= self.workbook_active.split('.')[0]
        # get data from scheme.json and stored it as dictionary
        dic = {}
        with open('scheme.json') as f:
            data = json.load(f)['classes']
            for d in data:
                if d['name'] == file_name:
                    dic = d
                    break
        development_pages = [d["development_goal"] for d in dic['development_page']]
        development_data = dic['development_page']
        # get the sheets in the workbook
        # start editing a pdf
        pdf = FPDF()
        print(dic)
        sheets = [sheet.title for sheet in self.workbook.worksheets]
        for sheet_name in sheets:
            if sheet_name in {'cover_page', 'first_page'}:
                sheet = self.workbook[sheet_name]
                fields = dic[sheet_name]
                attr = fields['report_fields']
                data = {'request': 'image'}
                headers = {'Authorization': f'Token {get_stored_token()}'}
                image_url = fields['page_background']
                new_url = url + image_url
                response = requests.get(new_url, data=data, headers=headers)

                values = self.get_values_from_sheet(sheet, self.selected_row)
                print(values)
                # get image from the response
                image = response.content
                with open('resources/' + sheet_name + '.jpg', 'wb') as f:
                    # save the image
                    f.write(image)
                # set image as pdf background
                pdf.add_page()
                # set the image to fit exactly the page
                pdf.image('resources/' + sheet_name + '.jpg', x=0, y=0, w=210, h=297)
                # set the font
                pdf.set_font("helvetica", size=17)
                # set the position of the text
                for field in attr:
                    if field['type_of_field'] != 'image':
                        # set the position of the text
                        pdf.set_xy(field['position_x'], field['position_y'])
                        # add the text to the pdf from values if its not none else add empty string
                        if field['multi_field']:
                            pdf.multi_cell(field['width'] if field['width'] else 200, field['height'] if field['height'] else 10,
                                           values[field['name']] if values[field['name']] is not None else "")
                        else:
                            pdf.cell(field['width'] if field['width'] else 200, field['height'] if field['height'] else 10,
                                 values[field['name']] if values[field['name']] is not None else "")
                    else:
                        # set the position of the image
                        pdf.set_xy(field['position_x'], field['position_y'])
                        # add the image to the pdf
                        pdf.image('resources/' + sheet_name + '.jpg', x=field['position_x'], y=field['position_y'],
                                  w=field['width'], h=field['height'])
            elif sheet_name == 'Image_page':
                sheet = self.workbook[sheet_name]
                fields = dic[sheet_name]
                values = self.get_values_from_sheet(sheet,self.selected_row)
                images=fields['images']
                print(images)
                new_url = url+fields["image"]
                headers = {'Authorization': f'Token {get_stored_token()}'}
                response = requests.get(new_url,  headers=headers)
                image = response.content
                with open('resources/image_background.jpg', 'wb') as f:
                    f.write(image)
                pdf.add_page()
                pdf.image('resources/image_background.jpg', x=0, y=0, w=210, h=297)
                for image in images:
                    pdf.image(values[image['title']],x=image['x'],y=image['y'],w=image['width'],h=image['height'])
                    print(values[image['title']])
                pdf.set_font("helvetica", size=17)
                print(fields)
            elif sheet_name == 'feedback_page':
                sheet = self.workbook[sheet_name]
                fields = dic[sheet_name]
                values = self.get_values_from_sheet(sheet, self.selected_row)
                sections = fields['sections']
                new_url = url+dic["default_background"]
                headers = {'Authorization': f'Token {get_stored_token()}'}
                response = requests.get(new_url,  headers=headers)
                image = response.content
                with open('resources/default_background.jpg', 'wb') as f:
                    f.write(image)
                pdf.add_page()
                pdf.image('resources/default_background.jpg', x=0, y=0, w=210, h=297)
                pdf.set_font("helvetica", size=22)
                current_y=21
                pdf.set_xy(21, current_y)
                pdf.cell(170, 10, 'Feedback', align='C')
                current_y+=17
                for section in sections:
                    pdf.set_font("helvetica", size=17)
                    pdf.set_xy(21, current_y)
                    pdf.cell(170, 12,section['name'], border=1, align='C')
                    current_y+=12
                    for field in section['Fields']:
                        pdf.set_font("helvetica", size=17)
                        pdf.set_xy(21, current_y)
                        # Calculate the height of the cell for the field name
                        name_no_of_cell= get_height(pdf,field['name'],80)
                        # Calculate the height of the cell for the field value
                        value_no_of_cell = get_height(pdf,values[field['name']],90)
                        print(name_no_of_cell, value_no_of_cell)
                        # Output the field name and value in multi_cell
                        if name_no_of_cell<value_no_of_cell:
                            cell_height = value_no_of_cell*10//name_no_of_cell
                            pdf.multi_cell(80, cell_height, field['name'], border=1, align='C')
                            pdf.set_xy(101, current_y)
                            pdf.multi_cell(90, 10, values[field['name']], border=1, align='C')
                            current_y+=value_no_of_cell*10
                        else:
                            cell_height = name_no_of_cell*10//value_no_of_cell
                            pdf.multi_cell(80, 10, field['name'], border=1, align='C')
                            pdf.set_xy(101, current_y)
                            pdf.multi_cell(90, cell_height, values[field['name']], border=1, align='C')
                            current_y+=name_no_of_cell*10
                    current_y+=5

            elif sheet_name in development_pages:
                # get the development page
                sheet = self.workbook[sheet_name]
                values = self.get_values_from_sheet_development(sheet, self.selected_row)
                # print(values)
                dev_page = development_data[development_pages.index(sheet_name)]
                new_url = url+dic["default_background"]
                headers = {'Authorization': f'Token {get_stored_token()}'}
                response = requests.get(new_url,  headers=headers)
                image = response.content
                with open('resources/default_background.jpg', 'wb') as f:
                    f.write(image)
                pdf.add_page()
                pdf.image('resources/default_background.jpg', x=0, y=0, w=210, h=297)
                pdf.set_font("helvetica", size=17)
                #create a table for the development page

                pdf.set_xy(21, 21)
                pdf.cell(170, 10, 'Development Goal:'+sheet_name, border=1, align='C')
                pdf.set_xy(21, 31)
                # multi cell for the key competencies and also get the number of cells that were added
                pdf.multi_cell(170, 10, 'Key Competencies:'+dev_page['key_components'], border=1, align='C',ln=1)
                # only set y position
                pdf.set_xy(21, pdf.get_y())
                pdf.cell(170/4, 20, '', border=1, align='C')
                pdf.set_xy(21+170/4, pdf.get_y())
                pdf.multi_cell(170/4, 10, 'Learning Outcome', border=1, align='C')
                pdf.set_xy(21+170/2, pdf.get_y()-20)
                pdf.cell(170/4, 20, 'Term 1', border=1, align='C')
                pdf.set_xy(21+3*170/4, pdf.get_y())
                pdf.cell(170/4, 20, 'Term 2', border=1, align='C')


                # get no of rows in each section
                section_rows_count = [ len(section['learning_outcome']) for section in dev_page['sections']]
                # print(section_rows_count)
                pdf.set_xy(21, pdf.get_y()+20)

                #get  

                # get the learning outcomes
                for i,section in enumerate(dev_page['sections']):
                    update_row_height = False
                    y_pos=pdf.get_y()
                    no_of_rows= get_height(pdf,section['name'],170/4)
                    print(no_of_rows,section['name'],section_rows_count[i])
                    height = 10 if no_of_rows>=section_rows_count[i] else 10*section_rows_count[i]/no_of_rows
                    if no_of_rows>section_rows_count[i]:
                        update_row_height = True
                    pdf.multi_cell(170/4,height,section['name'].title(),border=1,align='C')
                    pdf.set_xy(21+170/4, y_pos)
                    for learning_outcome in section['learning_outcome']:
                        height = no_of_rows*10/section_rows_count[i] if update_row_height else 10
                        pdf.cell(170/4, height, learning_outcome['code'], border=1, align='C')
                        pdf.set_xy(21+170/2, pdf.get_y())
                        pdf.cell(170/4, height, "" if values[learning_outcome['code']]['term 1']==None else values[learning_outcome['code']]['term 1'], border=1, align='C')
                        pdf.set_xy(21+3*170/4, pdf.get_y())
                        pdf.cell(170/4, height, "" if values[learning_outcome['code']]['term 2']==None else values[learning_outcome['code']]['term 2'], border=1, align='C')
                        pdf.set_xy(21+170/4, pdf.get_y()+height)
                    
                    pdf.set_xy(21, pdf.get_y())

                # add border to the cell
                



        pdf.output("resources/" + file_name + ".pdf")
        

    def get_values_from_sheet(self,sheet, selected_row):
        values = {}
        col = 1
        while col <= sheet.max_column:
            values[sheet.cell(1, col).value] = sheet.cell(selected_row, col).value
            col += 1
        return values
    
    def get_values_from_sheet_development(self, sheet, selected_row):
        values = {}
        col = 1
        while col <= sheet.max_column:
            header = sheet.cell(1, col).value
            stripped_header = header[:-7]
            if header.endswith('term 1'):
                if stripped_header in values:
                    values[stripped_header].update({'term 1': sheet.cell(selected_row, col).value})
                else:
                    values[stripped_header] = {'term 1': sheet.cell(selected_row, col).value}
            elif header.endswith('term 2'):
                if stripped_header in values:
                    values[stripped_header].update({'term 2': sheet.cell(selected_row, col).value})
                else:
                    values[stripped_header] = {'term 2': sheet.cell(selected_row, col).value}
            else:
                values[stripped_header] = sheet.cell(selected_row, col).value
            col += 1
        return values



class AddScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.file_manager = ImageManager()
        self.newImage=None
        self.development_page_count = 0
        self.section_no=''

    def on_enter(self, *args):
        self.values = []
        global url
        self.url = url
        self.development_page_count = 0
        self.ids.container.clear_widgets()
        self.get_xlsx_files()

        

    def get_xlsx_files(self):
        files = glob.glob('resources/*.xlsx')
        dropdown_items = [os.path.basename(f) for f in files]
        self.add_buttons(dropdown_items)

    def add_buttons(self, files):
        widge = Widget()
        self.ids.container.add_widget(widge)
        for file in files:
            button = MDRaisedButton(text=file, on_release=self.on_button_click, pos_hint={"center_x": 0.5})
            self.ids.container.add_widget(button)
        widge = Widget()
        self.ids.container.add_widget(widge)

    def on_button_click(self, instance):
        # remover everything inside the container
        self.ids.container.clear_widgets()
        # Create a ScrollView
        scroll_view = ScrollView(do_scroll_x=False, size_hint=(None, None), size=(400, 450), pos_hint={"center_x": 0.5})
        # Create a BoxLayout inside the ScrollView
        box_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        box_layout.bind(minimum_height=box_layout.setter('height'))
        scroll_view.add_widget(box_layout)

        # open the file name instance.text
        self.workbook_active = instance.text
        workbook = load_workbook(filename=f'resources/{instance.text}')

        # check if the user has access to create content from the file
        with open('scheme.json') as f:
            data = json.load(f)
            for d in data['classes']:
                if d['name'] == self.workbook_active.split('.')[0]:
                    data = d
                    break
            # print("data",data["cover_page_access"])
            if not (data["cover_page_access"]['username']==get_username() or get_username() == 'admin'):
                toast('You do not have access to this section')
                # close add screen
                self.manager.current = 'home'

                
        # get the first sheet
        sheets = [sheet.title for sheet in workbook.worksheets]

        # Create a list to hold the sheets
        self.sheets = sheets
        self.workbook = workbook
        self.current_sheet_index = 0

        # Create MDTextField for each cell in the first row of the first sheet
        self.create_text_fields(box_layout, self.current_sheet_index)
        

        # Add the ScrollView to the container
        self.ids.container.add_widget(Widget())
        self.ids.container.add_widget(scroll_view)
        self.ids.container.add_widget(Widget())

    def create_button_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]
        val = []

        # Create a button for each cell in the row
        for i in range(0, len(row), 2):
            cell1 = row[i]
            cell2 = row[i+1]
            if cell1.value and cell2.value:
                common_text = get_common_text(cell1.value, cell2.value)
                component = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 60))
                label = MDLabel(text=common_text)
                button1 = MDRaisedButton(text='Term 1', on_press=lambda x: change_text_field(x),id=common_text+'1')
                button2 = MDRaisedButton(text='Term 2',on_press=lambda x: change_text_field(x),id=common_text+'2')
                component.add_widget(label)
                component.add_widget(button1)
                component.add_widget(button2)
                box_layout.add_widget(component)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def create_feedback_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]

        # Create a button for each cell in the row
        for i in range(0, len(row)):
                component = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 60))
                label = MDLabel(text=row[i].value)
                # a button with dropdown arrow on the right

                button1 = MDRectangleFlatIconButton(text="None",icon="menu-down", on_press=lambda x:self.show_confirmation_dialog(x),id=row[i].value)
                component.add_widget(label)
                component.add_widget(button1)
                box_layout.add_widget(component)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def create_text_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]
        val = []

        # Create a MDTextField for each cell in the row
        for cell in row:
            if cell.value:
                text_field = MDTextField(hint_text=str(cell.value),id=str(cell.value).strip().replace(' ',"").lower())
                box_layout.add_widget(text_field)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)


    def show_confirmation_dialog(self,instance):
        feedbacks_options=[]
        # Add indentation here
        # to fix the "Expected indented block" error
        with open('scheme.json') as f:
            data = json.load(f)
            for d in data['classes']:
                if d['name'] == self.workbook_active.split('.')[0]:
                    data = d
                    break
            data=data['feedback_page']['sections']
            for d in data:
                for d1 in d['Fields']:
                    if d1['name'] == instance.id:
                        feedbacks_options = d1["options"]
                        break

        item=[ItemConfirm(text=i['choice']) for i in feedbacks_options]
        item.append(ItemConfirm(text='Other'))
        # print(feedbacks_options)
        self.dialog = MDDialog(
            title=instance.id,
            # feedbacks=feedbacks,
            type="confirmation",
            items=item,
            buttons=[
                MDFlatButton(
                    text="CANCEL",
                    theme_text_color="Custom",
                    on_release=self.close_dialog,
                ),
                MDFlatButton(
                    text="OK",
                    theme_text_color="Custom",
                    on_release=lambda x:self.close_dialog(x, True,instance),
                ),
            ],
        )
        self.dialog.open()

    def close_dialog(self, inst, update=False,instance=None):
        if update:
            val=[i.text for i in self.dialog.items if i.ids.check.active]
            # updated the text of button that has colled the function
            if len(val)>0:
                instance.text=val[0]
        self.dialog.dismiss()

    def create_image_fields(self, box_layout, sheet_index):
        # Clear the BoxLayout
        box_layout.clear_widgets()
        self.ids.container.clear_widgets()
        # Get the current sheet
        sheet = self.sheets[sheet_index]
        worksheet = self.workbook[sheet]

        # Get the first row
        row = worksheet[1]
        val = []

        # Create a MDTextField for each cell in the row
        for cell in row:
            if cell.value:
                component = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 60))
                label = MDLabel(text=cell.value)
                # a button with dropdown arrow on the right
                button1 = MDRectangleFlatIconButton(text="None",icon="folder",id=cell.value)
                button1.bind(on_press=lambda x,button1=button1:self.file_manager.file_manager_open(button1))
                component.add_widget(label)
                component.add_widget(button1)
                box_layout.add_widget(component)

        next_button = MDRaisedButton(text='Next', on_release=self.on_next_button_click, pos_hint={"center_x": 0.5})
        box_layout.add_widget(next_button)

    def get_admission_number_column(self, sheet):
        """
        Finds the column index containing "Admission Number" in the sheet's first row.

        Args:
            sheet: The sheet object from your workbook.

        Returns:
            The column index (integer) containing "Admission Number" or None if not found.
        """
        # Get the first row
        first_row = [cell.value for cell in sheet[1]]
        # print(first_row)
        # Find the index of the cell containing "Admission Number" (case-insensitive)
        for col_index, cell in enumerate(first_row):
            if cell and cell.lower() == "admission number":
                return col_index

        return None

    def on_next_button_click(self, instance):
        # iterate throught the text fields and get the values
        value_li = []
        sheet = self.sheets[self.current_sheet_index]
        if sheet in ['cover_page','first_page']:
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for text_field in widget.children:
                                if isinstance(text_field, MDTextField):
                                    print(text_field.text)
                                    if text_field.hint_text.strip()=='Admission Number':
                                        admission_number_column=self.get_admission_number_column(self.workbook[sheet])
                                        # print("admission number col: ",admission_number_column)
                                        if admission_number_column is not None:
                                            column = [ value[admission_number_column]  for value in self.workbook[sheet]]
                                            print("column: ",column)
                                            # skip the first row
                                            for cell in column[1:]:
                                                if cell.value == text_field.text.strip():
                                                    toast('Admission Number already exists')
                                                    self.values = []
                                                    return 
                                    value_li.append(text_field.text)
        elif self.sheets[self.current_sheet_index] in ['feedback_page']:
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for boxs in widget.children:
                                for button in boxs.children:
                                    if isinstance(button, MDRectangleFlatIconButton):
                                        value_li.append(button.text)
        elif self.sheets[self.current_sheet_index] in ['Image_page']:
            # Check if the images folder exists
            if not os.path.exists('resources/images'):
                os.makedirs('resources/images')
            if not os.path.exists(f'resources/images/{self.values[1][1]}'):
                os.makedirs(f'resources/images/{self.values[1][1]}')
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    i=0
                    self.newImage = SaveImage()
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for boxs in  widget.children:
                                files=['Me','Family']
                                for button in boxs.children:
                                    if isinstance(button, MDRectangleFlatIconButton):
                                            # print(i,button.text)
                                            ext=button.text.split('.')[-1]
                                            # print("ext: ",ext)
                                            value_li.append(f'resources/images/{self.values[1][1]}/{files[i]}.{ext}')
                                            self.newImage.addImage(i,button.text,f'resources/images/{self.values[1][1]}/{files[i]}.{ext}')
                                            # print(files[i])
                                            i+=1
        else:
            for child in self.ids.container.children:
                if isinstance(child, ScrollView):
                    for widget in child.children:
                        if isinstance(widget, BoxLayout):
                            for boxs in widget.children:
                                for button in boxs.children:
                                    if isinstance(button, MDRaisedButton):
                                        if button.text == 'Term 1' or button.text == 'Term 2':
                                            value_li.append("")
                                        else:
                                            value_li.append(button.text)
        
        self.current_sheet_index += 1
        self.values.append(value_li[::-1])

        # Clear the container
        self.ids.container.clear_widgets()
        # Create a ScrollView
        scroll_view = ScrollView(do_scroll_x=False, size_hint=(None, None), size=(400, 450), pos_hint={"center_x": 0.5})
        # Create a BoxLayout inside the ScrollView
        box_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        box_layout.bind(minimum_height=box_layout.setter('height'))
        scroll_view.add_widget(box_layout)

        click_next_button=False

        # If there are more sheets, create text fields for the next sheet
        if self.current_sheet_index < len(self.sheets):
            with open('scheme.json') as f:
                data = json.load(f)
                for d in data['classes']:
                    if d['name'] == self.workbook_active.split('.')[0]:
                        data = d
                        break
                else:
                    # notify that there was an error and returm to the main screen
                    toast('An error occured please contact the developer! This is related to your access to this section')
                    self.manager.current = 'home'
            if self.sheets[self.current_sheet_index] in ['cover_page','first_page']:
                access_accounts=data[self.sheets[self.current_sheet_index]+"_access"]
                self.create_text_fields(box_layout, self.current_sheet_index)
                # wait until the above call has completed
                if not (access_accounts['username']==get_username() or get_username() == 'admin'):
                    click_next_button=True
                else:
                    self.section_no+=','+str(self.current_sheet_index)
                # print("access accounts",access_accounts)
            elif self.sheets[self.current_sheet_index] in ['Image_page']:
                access_accounts=data[self.sheets[self.current_sheet_index]+"_access"]
                # print("access accounts",access_accounts)
                self.create_image_fields(box_layout, self.current_sheet_index)  
                if not (access_accounts['username']==get_username() or get_username() == 'admin'):
                    click_next_button=True
                else:
                    self.section_no+=','+str(self.current_sheet_index)
            elif self.sheets[self.current_sheet_index] in ['feedback_page']:
                self.section_no+=','+str(self.current_sheet_index)
                self.create_feedback_fields(box_layout, self.current_sheet_index)
            else:
                self.create_button_fields(box_layout, self.current_sheet_index)
                access_accounts=[dic['username'] for dic in data['development_page_access'][self.development_page_count]['Auth_teachers_access']]
                print("access accounts",access_accounts)
                self.development_page_count += 1
                if not (get_username() in access_accounts or get_username() == 'admin'):
                    click_next_button=True
                else:
                    self.section_no+=','+str(self.current_sheet_index)
            self.ids.container.add_widget(Widget())
            self.ids.container.add_widget(scroll_view)
            self.ids.container.add_widget(Widget())
            if click_next_button:
                self.on_next_button_click(instance)
        else:
            #display its done and add a home button
            self.ids.container.clear_widgets()
            self.ids.container.add_widget(Widget())
            self.ids.container.add_widget(MDLabel(text='Done', halign='center', theme_text_color='Primary'))
            self.ids.container.add_widget(Widget())
            sheet = self.workbook['cover_page']
            next_empty_row = sheet.max_row + 1
            for sheet, value in zip(self.sheets, self.values):
                worksheet = self.workbook[sheet]
                for i, value in enumerate(value, start=1):
                    worksheet.cell(row=next_empty_row, column=i, value=value)
            if self.newImage or self.newImage=='None':
                self.newImage.save_image()
            self.workbook.save(f'resources/{self.workbook_active}')
            url = self.url + "update/"
            print("value:",self.values[1])
            data = {"request": "update","section_no":'all','admission_no':self.values[1][1],'class_name':self.workbook_active.split('.')[0]}
            headers = {'Authorization': f'Token {get_stored_token()}'}
            response = requests.post(url, headers=headers, data=data)
            if response.status_code == 200:
                toast('Server notified successfully\n Your commit number is: '+str(response.json()['commit_no']))
                # open file user.json
                with open('user.json') as f:
                    user = json.load(f)
                    last_updated_commit_no = user.get("commit_no",0)
                    if last_updated_commit_no==0:
                        user['commit_no']={self.workbook_active.split('.')[0]:0}
                        last_updated_commit_no = 0
                    else:
                        last_updated_commit_no = last_updated_commit_no.get(self.workbook_active.split('.')[0],0)
                    print(last_updated_commit_no,response.json()['commit_no'])
                    if last_updated_commit_no + 1 == response.json()['commit_no']:
                        user['commit_no'][self.workbook_active.split('.')[0]]=response.json()['commit_no']
                        # update the user in the user.json file
                        with open('user.json', 'w') as f:
                            json.dump(user, f)
                print(response.json()['devices'])
                for device in response.json()['devices']:
                    # check that device ip is not loop back ip
                    ip_address=get_global_ipv6_address()
                    if not device['device_ip'].startswith('fe80') and not device['device_ip'].startswith('fd') and device['device_ip'] != '::1' and device['device_ip'] != ip_address:
                        sheets = [sheet.title for sheet in self.workbook.worksheets]
                        section_no=self.section_no.strip(',')
                        section_no=list(map(int,section_no.split(',')))
                        admission_no=self.values[1][1]
                        sheet=self.workbook['cover_page']
                        #in first row find the cell with value admission number
                        match=False
                        for cell in sheet[1]:
                            if cell.value == 'Admission Number':
                                row:int=2
                                while row <= sheet.max_row:
                                    if str(sheet.cell(row=row, column=cell.column).value).strip() == str(admission_no).strip():
                                        match=True
                                        selected_row = row
                                    row += 1
                        result= {}
                        files={}
                        if match:
                            for section in section_no:
                                for a in self.workbook[sheets[section]][selected_row]:
                                    if a.value and check_if_path(a.value):
                                        # result[section]='file_to_share'+str(file_count)
                                        files[a.value] = encode_image_to_base64(a.value)
                                result[section]=[a.value for a in self.workbook[sheets[section]][selected_row]]
                            print("results:",result)
                            # open file user.json
                            with open('user.json') as f:
                                user = json.load(f)
                                last_updated_commit_no = user.get("commit_no",0)
                                if last_updated_commit_no==0:
                                    user['commit_no']={self.workbook_active.split('.')[0]:0}
                                    last_updated_commit_no = 0
                                else:
                                    last_updated_commit_no = last_updated_commit_no.get(self.workbook_active.split('.')[0],0)
                                if last_updated_commit_no + 1 == response.json()['commit_no']:
                                    user['commit_no'][self.workbook_active.split('.'[0])]=response.json()['commit_no']
                                    # update the user in the user.json file
                                    with open('user.json', 'w') as f:
                                        json.dump(user, f)
                        json_data={"message":"Initiating commit push",
                                   "commit_no":str(response.json()['commit_no']),
                                   "admission_no":self.values[1][1],
                                   "class_name":self.workbook_active.split('.')[0],
                                   "section_no":self.section_no,
                                   "results":result,
                                   'files': files}
                        
                        with open('data.json','w') as f:
                            json.dump(json_data,f)

                        connect_to_server_thread(device['device_ip'],1680,json_data)
            else:
                # create a file named notification.txt
                with open('notification.txt', 'a') as f:
                    f.write("\n"+self.section_no+":"+self.values[1][1]+":"+self.workbook_active.split('.')[0])
                toast('An error occured while notifying the server')



class MainApp(MDApp):
    def __init__(self, **kwargs):
        super().__init__()

    def build(self):
        global url
        self.url = url
        self.sm = ScreenManager()
        # check if there is a resources folder
        if not os.path.exists('resources'):
            os.makedirs('resources')
        self.sm.add_widget(LoginScreen(name='login'))
        self.sm.add_widget(HomeScreen(name='home'))
        self.sm.add_widget(EditScreen(name='edit'))
        self.sm.add_widget(PrintScreen(name='print'))
        self.sm.add_widget(AddScreen(name='add'))

        # Create or connect to the SQLite database
        # self.conn = sqlite3.connect('mydatabase.db')
        # self.cursor = self.conn.cursor()

        # # Create a table if it doesn't exist
        # self.cursor.execute('''CREATE TABLE IF NOT EXISTS Person (
        #                                id INTEGER PRIMARY KEY,
        #                                name TEXT,
        #                                age INTEGER)''')
        # self.cursor.execute('''CREATE TABLE IF NOT EXISTS Cookies (
        #                         Name TEXT PRIMARY KEY,
        #                         Value TEXT
        #     )''')
        # self.conn.commit()
        data = {"request": "access"}
        # cookies = {"sessionid": get_sessionid(self.cursor)}
        headers = {'Authorization': f'Token {get_stored_token()}'}
        response = requests.post(url + "data/", headers=headers, data=data)
        # response = requests.post(url, data=data, cookies=cookies)
        if response.status_code == 200:
            self.sm.current = 'home'
        self.stop_event = threading.Event()
        server_thread = threading.Thread(target=start_server, args=(get_global_ipv6_address(), 1680, self.stop_event))
        server_thread.start()
        return self.sm
    
    def show_confirmation_dialog(self):
        if not self.dialog:
            self.dialog = MDDialog(
                title="Phone ringtone",
                type="confirmation",
                items=[
                    ItemConfirm(text="Callisto"),
                    ItemConfirm(text="Luna"),
                    ItemConfirm(text="Night"),
                    ItemConfirm(text="Solo"),
                    ItemConfirm(text="Phobos"),
                    ItemConfirm(text="Diamond"),
                    ItemConfirm(text="Sirena"),
                    ItemConfirm(text="Red music"),
                    ItemConfirm(text="Allergio"),
                    ItemConfirm(text="Magic"),
                    ItemConfirm(text="Tic-tac"),
                    ItemConfirm(text="Other"),
                ],
                buttons=[
                    MDFlatButton(
                        text="CANCEL",
                        theme_text_color="Custom",
                        text_color=self.theme_cls.primary_color,
                        on_release=self.close_dialog,
                    ),
                    MDFlatButton(
                        text="OK",
                        theme_text_color="Custom",
                        text_color=self.theme_cls.primary_color,
                        on_release=lambda x:self.close_dialog(x, True),
                    ),
                ],
            )
        self.dialog.open()

    def close_dialog(self, inst, update=False):
        if update:
            print("Update", [i.text for i in self.dialog.items if i.ids.check.active])
        self.dialog.dismiss()

    def on_login(self, username, password):
        # API endpoint
        url = self.url + "login/"

        # Data to be sent to the API
        data = {"username": username, "password": password}
        # Send a post request to the API
        response = requests.post(url, data=data)
        if response.status_code == 200:
            # create a file make user.json and store the username
            with open('user.json', 'w') as f:
                json.dump({'username': username}, f)
            # Assuming the token is returned in the response JSON
            response_json = response.json()
            new_token = response_json.get('token')
            if new_token:
                store_token(new_token)
            else:
                print("Token not found in the response.")
        else:
            print("Login failed. Status code:", response.status_code)

        # if 'sessionid' in response.cookies:
        #     sessionid = response.cookies['sessionid']
        #     self.cursor.execute("INSERT OR REPLACE INTO Cookies (Name, Value) VALUES (?, ?)", ("sessionid", sessionid))

        # If the response status code is 200 (HTTP OK), switch to the home screen
        if response.status_code == 200:
            # url = "http://127.0.0.1:8000/data/"
            # cookies = {"sessionid": get_sessionid(self.cursor)}
            # data = {"request":"schema"}
            self.sm.current = 'home'
        else:
            snackbar = MDSnackbar(
                MDLabel(
                    text="Invalid Username or Password",
                ),
                MDSnackbarActionButton(
                    text="DISMISS",
                    on_release=lambda *args: snackbar.dismiss(),
                    theme_text_color="Custom",
                    text_color="#8E353C",
                ),
                y=dp(24),
                pos_hint={"center_x": 0.5},
                size_hint_x=0.5,
                md_bg_color="#E8D8D7",
            )
            snackbar.open()

    def logout(self):
        self.sm.current = 'login'
        #     clear the username and password fields
        url = self.url + "data/"
        data = {"request": "logout"}
        # cookies = {"sessionid": get_sessionid(self.cursor)}
        headers = {'Authorization': f"Token {get_stored_token()}"}
        response = requests.post(url, headers=headers, data=data)
        if response.status_code == 200:
            store_token("")
            # clear the user.json file
            with open('user.json', 'w') as f:
                json.dump({}, f)
        self.sm.get_screen('login').ids['username'].text = ''
        self.sm.get_screen('login').ids['password'].text = ''

    def go_home(self):
        self.sm.current = 'home'

    def on_start(self):
        ImLive(0)  # call ImLive once to start the schedule
        Clock.schedule_interval(ImLive, 150)  # schedule ImLive to be called every 30 seconds
        return super().on_start()
    
    def on_stop(self):
        self.stop_event.set()


    
MainApp().run()
