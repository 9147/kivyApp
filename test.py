import json
import os
import logging
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from dependant import check_if_path
from imageConversion import decode_base64_to_image

def process_commit_push(received_data):
        wb = load_workbook("resources/" + received_data.get('class_name') + '.xlsx')
        sheets = [sheet.title for sheet in wb.worksheets]
        section_no = received_data.get('section_no').strip(',').split(',')
        section_no = list(map(int, section_no))
        admission_no = received_data.get('admission_no')
        sheet = wb['cover_page']

        match = False
        for cell in sheet[1]:
            if cell.value == 'Admission Number':
                row = 2
                while row <= sheet.max_row:
                    if str(sheet.cell(row=row, column=cell.column).value).strip() == str(admission_no).strip():
                        match = True
                        selected_row = row
                    row += 1
        
        # check if there is recources folder
        if not os.path.exists('resources'):
            os.makedirs('resources')
        if not os.path.exists('resources/images'):
            os.makedirs('resources/images')
        if not os.path.exists(f'resources/images/{admission_no}'):
                os.makedirs(f'resources/images/{admission_no}')
        file=received_data['files']
        if match:
            for section in section_no:
                sheet = wb[sheets[section]]
                row = 0
                for cell in sheet[selected_row]:
                    print(row,received_data.get('results').get(str(section)),section)
                    val = received_data.get('results').get(str(section))[row]
                    if check_if_path(val):
                        decode_base64_to_image(file[val],val)
                    else:
                        cell.value = val 
                    row += 1
        else:
            logging.info("Admission number not found")
            sheet = wb['cover_page']
            next_empty_row = sheet.max_row + 1
            for section in section_no:
                worksheet = wb[sheets[section]]
                values = received_data.get('results').get(str(section))
                for i, value in enumerate(values, start=1):
                    for a in value:
                        if check_if_path(a):
                            decode_base64_to_image(file[a],a)
                    worksheet.cell(row=next_empty_row, column=i, value=value)
        wb.save("resources/" + received_data.get('class_name') + '.xlsx')



with open('data.json','r') as f:
    data = json.load(f)
    # print(data)
    process_commit_push(data)
    