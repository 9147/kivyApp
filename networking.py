import socket
import netifaces
import json
import random
import os
import pickle
import string
from openpyxl import load_workbook
import logging
from imageConversion import decode_base64_to_image
from dependant import check_if_path
import threading

buffer_size=1024*1024
CHUNK_SIZE = 1024  # 1 KB
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def get_global_ipv6_address():
    interfaces = netifaces.interfaces()
    for interface in interfaces:
        addresses = netifaces.ifaddresses(interface)
        if netifaces.AF_INET6 in addresses:
            for addr in addresses[netifaces.AF_INET6]:
                ipv6_addr = addr['addr']
                # Check for a global unicast address (not starting with fe80:: or fd00::/8)
                if ipv6_addr and not ipv6_addr.startswith('fe80') and not ipv6_addr.startswith('fd'):
                    return ipv6_addr.split('%')[0]  # Remove the zone index if present
    return None



def handle_client(conn, addr):
    logging.info(f"Connected by {addr}")
    try:
        # Receive data in chunks
        received_data_bytes = b""
        while True:
            chunk = conn.recv(CHUNK_SIZE)
            if not chunk:
                break
            received_data_bytes += chunk

        received_data = json.loads(received_data_bytes.decode('utf-8'))
        # logging.info(f"Received: {received_data}")

        message = received_data.get("message")
        # logging.info(f"Message: {message}")
        # logging.info(f"Initiating commit push: {message == 'Initiating commit push'}")
        if message == 'Initiating commit push':
            with open('data.json', 'w') as f:
                json.dump(received_data, f)
            process_commit_push(received_data)
            # open file user.json
            with open('user.json') as f:
                user = json.load(f)
                last_updated_commit_no = user.get("commit_no", 0)
                if last_updated_commit_no == 0:
                    user['commit_no'] = {received_data['class_name']: 0}
                    last_updated_commit_no = 0
                else:
                    last_updated_commit_no = last_updated_commit_no.get(received_data['class_name'], 0)
                if last_updated_commit_no + 1 == received_data['commit_no']:
                    user['commit_no'][received_data['class_name']] = received_data['commit_no']
                    # update the user in the user.json file
                    with open('user.json', 'w') as f:
                        json.dump(user, f)
            response = json.dumps({"message": "Commit push initiated"})
            conn.sendall(response.encode('utf-8'))
        else:
            response = json.dumps({"message": "Hello from the server!"})
            conn.sendall(response.encode('utf-8'))
    except Exception as e:
        logging.error(f"Error processing data: {e}")
    finally:
        conn.close()

def start_server(ipv6_address, port, stop_event):
    server_socket = socket.socket(socket.AF_INET6, socket.SOCK_STREAM)
    server_socket.bind((ipv6_address, port, 0, 0))
    server_socket.listen(5)  # Allow up to 5 queued connections
    server_socket.settimeout(1)  # set a timeout of 1 second
    logging.info(f"Server listening on [{ipv6_address}]:{port}")

    while not stop_event.is_set():
        try:
            conn, addr = server_socket.accept()
            client_thread = threading.Thread(target=handle_client, args=(conn, addr))
            client_thread.start()
        except socket.timeout:
            continue  # if a timeout occurs, continue the loop to check the stop_event
        except Exception as e:
            logging.error(f"Error accepting connection: {e}")
            continue

    server_socket.close()

def process_commit_push(received_data):
    try:
        wb = load_workbook("resources/" + received_data.get('class_name') + '.xlsx')
        sheets = [sheet.title for sheet in wb.worksheets]
        section_no = received_data.get('section_no').strip(',').split(',')
        section_no = list(map(int, section_no))
        admission_no = received_data.get('admission_no')
        sheet = wb['cover_page']

        match = False
        for cell in sheet[1]:
            if cell.value == 'Admission Number':
                row = 2
                while row <= sheet.max_row:
                    if str(sheet.cell(row=row, column=cell.column).value).strip() == str(admission_no).strip():
                        match = True
                        selected_row = row
                    row += 1
        
        # check if there is recources folder
        if not os.path.exists('resources'):
            os.makedirs('resources')
        if not os.path.exists('resources/images'):
            os.makedirs('resources/images')
        if not os.path.exists(f'resources/images/{admission_no}'):
                os.makedirs(f'resources/images/{admission_no}')
        file=received_data['files']
        if match:
            for section in section_no:
                sheet = wb[sheets[section]]
                row = 0
                for cell in sheet[selected_row]:
                    val = received_data.get('results').get(str(section))[row]
                    if val and os.path.exists('/'.join(val.split('/')[:-1])):
                        decode_base64_to_image(file[val],val)
                    else:
                        cell.value = val 
                    row += 1
        else:
            logging.info("New Entry detected!!")
            sheet = wb['cover_page']
            next_empty_row = sheet.max_row + 1
            for section in section_no:
                worksheet = wb[sheets[section]]
                values = received_data.get('results').get(str(section))
                for i, value in enumerate(values, start=1):
                    if value and os.path.exists('/'.join(value.split('/')[:-1])):
                        decode_base64_to_image(file[value],value)
                    worksheet.cell(row=next_empty_row, column=i, value=value)
        wb.save("resources/" + received_data.get('class_name') + '.xlsx')
    except Exception as e:
        logging.error(f"Error processing commit push: {e}")

def connect_to_server_thread(device_ip, port, message_dict, timeout=5):
    def connect_to_server(ipv6_address, port, message_dict, timeout):
        client_socket = socket.socket(socket.AF_INET6, socket.SOCK_STREAM)
        client_socket.settimeout(timeout)
        try:
            client_socket.connect((ipv6_address, port, 0, 0))
            logging.info(f"Connected to server at [{ipv6_address}]:{port}")

            message = json.dumps(message_dict)
            message_bytes = message.encode('utf-8')

            # Send data in chunks
            for i in range(0, len(message_bytes), CHUNK_SIZE):
                chunk = message_bytes[i:i + CHUNK_SIZE]
                client_socket.sendall(chunk)

            # Receive data in chunks
            received_data_bytes = b""
            while True:
                chunk = client_socket.recv(CHUNK_SIZE)
                if not chunk:
                    break
                received_data_bytes += chunk

            received_data = json.loads(received_data_bytes.decode('utf-8'))
            logging.info(f"Received from server: {received_data}")
        except socket.timeout:
            logging.error(f"Connection to [{ipv6_address}]:{port} timed out")
        except Exception as e:
            logging.error(f"Error connecting to server: {e}")
        finally:
            client_socket.close()

    thread = threading.Thread(target=connect_to_server, args=(device_ip, port, message_dict, timeout))
    thread.start()

        
    

def generate_code():
    # generate a random 6 alpha numbric code
    code = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
    # get or create a code.bin file that stores a set of codes
    if os.path.exists('code.bin'):
        with open('code.bin', 'rb') as file:
            codes = pickle.load(file)
    else:
        codes = set()
    codes.add(code)
    with open('code.bin', 'wb') as file:
        pickle.dump(codes, file)
    return code

if __name__ == "__main__":
    ipv6_address = get_global_ipv6_address()
    if ipv6_address:
        port = 1680
        start_server(ipv6_address, port)
    else:
        print("No global IPv6 address found.")
